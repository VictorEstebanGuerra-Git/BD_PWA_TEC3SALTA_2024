'''
Autor: Ramiro Sebastian Gaspar
Fecha: 19/06/2024
Proyecto: PWA ETT 3139 GRAL Martin Miguel de Guemes
Institucion: E.E.T 3139 Gral. M.M de Guemes, Salta Capital

Proposito: Este codigo crea un excel, con el fin de que el equipo de testeo no tenga que estar creando un excel
manualmente y que tenga datos incorrectos a los que ya son predeterminados en la bd.

Notas: IMPORTANTE, tienen que ejecutar la cantidad de veces que sea necesaria para rellenar los atributos que contengan
JSON como tipo de dato. Ejemplo, si quieren empezar a hacer pruebas de la tabla de notas tienen que ejecutar 4 veces
este codigo PERO cambiando el nombre en la direccion del aarchivo asi no se sobre escriben, como en el caso de la nota
al final de la direccion ponen trimestre1.xlsx, despues lo ejecutan de nuevo y ponen trimestre2.xlsx, y asi sucesivamente
perdon por realizarlo de esta forma, es por falta de tiempo.

Se tiene que instalar la siguiente libreria:
pip install openpyxl

''' 

# Se importan las librerias
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Crear un nuevo libro de trabajo
wb = Workbook()
ws = wb.active

# Definir el borde negro
border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

# Aplicar el borde y el estilo a las celdas A1:A13
for row in range(1, 14):
    cell = ws.cell(row=row, column=1, value="")
    cell.fill = PatternFill(start_color='B6D7A8', end_color='B6D7A8', fill_type='solid')
    cell.font = Font(bold=True)
    cell.alignment = Alignment(wrap_text=True)
    cell.border = border

# Agregar los elementos especificados
elementos = [
    "curso",
    "ciclo",
    "materia",
    "docente",
    "fecha",
    "dni alumno",
    "nombre alumno",
    "apellido alumno",
    "concepto",
    "nota",
    "estado",
    "firma profesor",
    "firma tutor"
]

# Escribir los elementos en las celdas correspondientes, una columna debajo de la otra
for row_index, value in enumerate(elementos, start=1):
    cell = ws.cell(row=row_index, column=1, value=value)
    cell.alignment = Alignment(wrap_text=True)

# Ajustar ancho de columna automáticamente
ws.column_dimensions['A'].bestFit = True
    
# Especificar la ruta donde se guardará el archivo, ACUERDENSE DE CAMBIAR LA PARTE FINAL
ruta_archivo = r"C:\Users\Usuario\Desktop\BD_PWA_ETT3139 Salta_2024\Excels\Recuperacion.xlsx"

# Guardar el archivo en la ruta especificada
wb.save(ruta_archivo)

print(f"Archivo guardado en: {ruta_archivo}")
